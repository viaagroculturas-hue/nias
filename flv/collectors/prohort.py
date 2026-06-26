"""
CONAB PROHORT — Preços de Hortifruti por CEASA
Fonte: Boletim Hortigranjeiro mensal CONAB/PROHORT
Cobre: 11 terminais CEASA em todo o Brasil
Produtos: Alface, Batata, Cebola, Cenoura, Tomate, Banana, Laranja, Maçã, Mamão, Melancia
"""
import re, io, time

_prohort_cache = {}
_prohort_ttl   = 6 * 3600  # 6h — boletim é mensal

# Mapeamento de nome CEASA (prefixo) → slug padronizado
CEASA_SLUG = {
    'CEAGESP':      'CEAGESP-SP',
    'CEASAMINAS':   'CEASA-MG',
    'CEASA/RJ':     'CEASA-RJ',
    'CEASA/SP':     'CEASA-SP-CAMPINAS',
    'CEASA/ES':     'CEASA-ES',
    'CEASA/PR':     'CEASA-PR',
    'CEASA/SC':     'CEASA-SC',
    'CEASA/GO':     'CEASA-GO',
    'CEASA/DF':     'CEASA-DF',
    'CEASA/PE':     'CEASA-PE',
    'CEASA/CE':     'CEASA-CE',
    'CEASA/BA':     'CEASA-BA',
    'CEASA/RS':     'CEASA-RS',
    'MEDIA':        '_nacional',
    'MÉDIA':        '_nacional',
}

CEASA_LABEL = {
    'CEAGESP-SP':          'CEAGESP — São Paulo/SP',
    'CEASA-MG':            'CEASAMINAS — Belo Horizonte/MG',
    'CEASA-RJ':            'CEASA/RJ — Rio de Janeiro/RJ',
    'CEASA-SP-CAMPINAS':   'CEASA/SP — Campinas/SP',
    'CEASA-ES':            'CEASA/ES — Vitória/ES',
    'CEASA-PR':            'CEASA/PR — Curitiba/PR',
    'CEASA-SC':            'CEASA/SC — São José/SC',
    'CEASA-GO':            'CEASA/GO — Goiânia/GO',
    'CEASA-DF':            'CEASA/DF — Brasília/DF',
    'CEASA-PE':            'CEASA/PE — Recife/PE',
    'CEASA-CE':            'CEASA/CE — Fortaleza/CE',
    'CEASA-BA':            'CEASA/BA — Salvador/BA',
    'CEASA-RS':            'CEASA/RS — Porto Alegre/RS',
}


def _slug_ceasa(raw: str) -> str:
    """Converte nome bruto do Excel para slug padronizado."""
    up = raw.upper().strip()
    for prefix, slug in CEASA_SLUG.items():
        if up.startswith(prefix.upper()):
            return slug
    return None


def _safe_float(v) -> float | None:
    if v is None:
        return None
    try:
        f = float(v)
        return None if f == 0 and str(v) in ('', None) else round(f, 4)
    except (ValueError, TypeError):
        return None


def _parse_price_sheet(ws, products: list[str]) -> dict:
    """
    Lê sheet de preços com estrutura padrão PROHORT:
      Row 0-3: cabeçalho (mês, título, data)
      Row 3: Produto   Prod1   ""    Prod2   ""  ...
      Row 4: Ceasa     Preço   Var%  Preço   Var%  ...
      Row 5+: CEASA    val     var   val     var  ...
      Última linha com dado: Média Ponderada
    Retorna {ceasa_slug: {produto_slug: {price, var_pct}}}
    """
    rows = list(ws.iter_rows(values_only=True))

    # Detectar linha de produtos (contém 'Produto' na primeira coluna)
    prod_row_idx = None
    for i, row in enumerate(rows):
        if row and str(row[0] or '').strip().upper() in ('PRODUTO', 'PRODUTOS'):
            prod_row_idx = i
            break
    if prod_row_idx is None:
        return {}

    prod_row  = rows[prod_row_idx]
    # header_row: linha seguinte (Ceasa / Preço / Var%)
    data_start = prod_row_idx + 2  # pula linha de header Ceasa/Preço/Var%

    # Mapear coluna index → produto
    col_map: dict[int, str] = {}   # col_index → produto_slug
    for ci, cell in enumerate(prod_row):
        if ci == 0 or not cell:
            continue
        cell_str = str(cell).strip()
        for pi, pname in enumerate(products):
            if pname.upper() in cell_str.upper():
                col_map[ci] = pname
                break

    result: dict = {}
    for row in rows[data_start:]:
        if not row or not row[0]:
            continue
        ceasa_raw = str(row[0]).strip()
        if ceasa_raw.upper().startswith('FONTE'):
            break
        slug = _slug_ceasa(ceasa_raw)
        if not slug:
            continue
        entry: dict = {}
        for ci, pslug in col_map.items():
            price = _safe_float(row[ci] if ci < len(row) else None)
            var   = _safe_float(row[ci + 1] if ci + 1 < len(row) else None)
            if price and price > 0:
                entry[pslug] = {'price_kg': price, 'var_pct': var}
        if entry:
            result[slug] = entry

    return result


def _find_latest_xlsx_url(session) -> str | None:
    """Navega no site CONAB para encontrar a URL do XLSX mais recente."""
    from bs4 import BeautifulSoup

    year = time.strftime('%Y')
    base = f'https://www.gov.br/conab/pt-br/atuacao/informacoes-agropecuarias/hortigranjeiros-prohort/boletim-hortigranjeiro/boletim-hortigranjeiro-{year}'

    try:
        r = session.get(base, timeout=20)
        if r.status_code != 200:
            # Tenta ano anterior
            year = str(int(year) - 1)
            base = base.replace(time.strftime('%Y'), year)
            r = session.get(base, timeout=20)
        if r.status_code != 200:
            return None

        soup = BeautifulSoup(r.text, 'html.parser')
        # Encontrar link do XLSX (tabelas-de-dados) — gov.br adiciona /view no final
        xlsx_links = []
        for a in soup.find_all('a', href=True):
            href = a['href']
            if 'tabelas-de-dados' in href.lower() and '.xlsx' in href.lower():
                # Remove /view suffix para obter URL de download direto
                clean = href.split('/view')[0]
                xlsx_links.append(clean)

        if not xlsx_links:
            return None

        # Ordena por mês usando posição do nome do mês na URL
        MONTHS = ['janeiro','fevereiro','marco','abril','maio','junho',
                  'julho','agosto','setembro','outubro','novembro','dezembro']
        def _month_order(url):
            ul = url.lower()
            for i, m in enumerate(MONTHS):
                if m in ul:
                    return i
            return -1

        xlsx_links.sort(key=_month_order)
        return xlsx_links[-1]  # mais recente

    except Exception as e:
        print(f'[PROHORT] Erro ao buscar URL: {e}')
        return None


def fetch_prohort() -> dict:
    """
    Retorna preços de hortifruti de 11 CEASAs brasileiras via CONAB PROHORT.
    Cache de 6 horas (boletim é atualizado mensalmente).
    """
    global _prohort_cache
    now = time.time()

    if _prohort_cache.get('data') and now - _prohort_cache.get('ts', 0) < _prohort_ttl:
        return _prohort_cache['data']

    try:
        import requests as rq
        import openpyxl
    except ImportError as e:
        print(f'[PROHORT] Dependência ausente: {e}')
        return {}

    try:
        from bs4 import BeautifulSoup
    except ImportError:
        print('[PROHORT] beautifulsoup4 não instalado')
        return {}

    s = rq.Session()
    s.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    })

    xlsx_url = _find_latest_xlsx_url(s)
    if not xlsx_url:
        print('[PROHORT] XLSX não encontrado')
        return {}

    print(f'[PROHORT] Baixando: {xlsx_url}')
    resp = s.get(xlsx_url, timeout=40)
    if resp.status_code != 200:
        print(f'[PROHORT] Download falhou: {resp.status_code}')
        return {}

    # Extrair mês/ano da URL
    month_match = re.search(
        r'(janeiro|fevereiro|marco|abril|maio|junho|julho|agosto|setembro|outubro|novembro|dezembro)-(\d{4})',
        xlsx_url.lower()
    )
    ref_period = month_match.group(0).replace('-', '/') if month_match else time.strftime('%m/%Y')

    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)

    # Parsear hortaliças
    horti_data: dict = {}
    if 'Preços-Hortaliças' in wb.sheetnames:
        horti_data = _parse_price_sheet(
            wb['Preços-Hortaliças'],
            ['Alface', 'Batata', 'Cebola', 'Cenoura', 'Tomate']
        )
    elif any('Hortali' in s for s in wb.sheetnames):
        ws_name = next(s for s in wb.sheetnames if 'Hortali' in s and 'Pre' in s)
        horti_data = _parse_price_sheet(
            wb[ws_name],
            ['Alface', 'Batata', 'Cebola', 'Cenoura', 'Tomate']
        )

    # Parsear frutas
    frutas_data: dict = {}
    if 'Preços-Frutas' in wb.sheetnames:
        frutas_data = _parse_price_sheet(
            wb['Preços-Frutas'],
            ['Banana', 'Laranja', 'Maçã', 'Mamão', 'Melancia']
        )
    elif any('Fruta' in s for s in wb.sheetnames):
        ws_name = next(s for s in wb.sheetnames if 'Fruta' in s and 'Pre' in s)
        frutas_data = _parse_price_sheet(
            wb[ws_name],
            ['Banana', 'Laranja', 'Maçã', 'Mamão', 'Melancia']
        )

    # Unificar por CEASA
    all_ceasas: set = set(horti_data.keys()) | set(frutas_data.keys())
    terminals: dict = {}
    for slug in all_ceasas:
        terminals[slug] = {
            **(horti_data.get(slug, {})),
            **(frutas_data.get(slug, {})),
            '_label': CEASA_LABEL.get(slug, slug),
        }

    # Nacional (média ponderada)
    nacional = {
        **(horti_data.get('_nacional', {})),
        **(frutas_data.get('_nacional', {})),
    }

    result = {
        'meta': {
            'source': 'CONAB/PROHORT',
            'period': ref_period,
            'xlsx_url': xlsx_url,
            'updated_at': time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime()),
            'terminals_count': len([k for k in terminals if not k.startswith('_')]),
            'products': ['Alface', 'Batata', 'Cebola', 'Cenoura', 'Tomate',
                         'Banana', 'Laranja', 'Maçã', 'Mamão', 'Melancia'],
        },
        'terminals': terminals,
        'nacional': nacional,
    }

    _prohort_cache['data'] = result
    _prohort_cache['ts'] = now
    print(f'[PROHORT] OK — {len(terminals)} terminais · {ref_period}')
    return result


if __name__ == '__main__':
    data = fetch_prohort()
    meta = data.get('meta', {})
    print(f"\nPROHORT — {meta.get('period','')} — {meta.get('terminals_count',0)} terminais")
    print(f"{'='*70}")
    for ceasa, prods in data.get('terminals', {}).items():
        if ceasa.startswith('_'):
            continue
        label = prods.get('_label', ceasa)
        print(f"\n  {label}")
        for prod, vals in prods.items():
            if prod.startswith('_'):
                continue
            if isinstance(vals, dict):
                p = vals.get('price_kg', 0)
                v = vals.get('var_pct', 0)
                sign = '+' if (v or 0) >= 0 else ''
                print(f"    {prod:12s}: R$ {p:6.2f}/kg  ({sign}{(v or 0)*100:.1f}% vs mês ant.)")
    print(f"\n--- MÉDIA NACIONAL ---")
    for prod, vals in data.get('nacional', {}).items():
        if isinstance(vals, dict):
            p = vals.get('price_kg', 0)
            print(f"  {prod:12s}: R$ {p:6.2f}/kg")
